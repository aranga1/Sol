"""ExcelExtractor — one TextNode per sheet using openpyxl.

Each row is serialised as "col1: val1 | col2: val2" where the column
headers come from the first row of the sheet.  If the sheet is empty,
it is omitted.
"""
from __future__ import annotations

from pathlib import Path

from llama_index.core.schema import TextNode


class ExcelExtractor:
    """Extract tabular data from an Excel workbook, one TextNode per sheet."""

    supported_extensions: frozenset[str] = frozenset({".xlsx", ".xls"})

    def extract(self, path: Path) -> list[TextNode]:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        base_meta = {"file_path": str(path), "file_name": path.name}
        nodes: list[TextNode] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First row treated as headers; fall back to column indices
            headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]

            lines: list[str] = []
            for row in rows:
                parts = []
                for header, cell in zip(headers, row):
                    val = "" if cell is None else str(cell)
                    parts.append(f"{header}: {val}")
                lines.append(" | ".join(parts))

            text = "\n".join(lines)
            meta = {**base_meta, "sheet_name": sheet_name}
            node = TextNode(text=text, metadata=meta)
            node.excluded_llm_metadata_keys = list(base_meta.keys())
            nodes.append(node)

        wb.close()
        return nodes
